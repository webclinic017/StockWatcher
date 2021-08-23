import formulas
import json
import re
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet import worksheet
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formula import Tokenizer
# CREATE WORKBOOK
# workbook = Workbook()
# sheet = workbook.active

# sheet["A1"] = "hello"
# sheet["B1"] = "world!"

# workbook.save(filename="../../../data/excel/data.xlsx")

# workbook = load_workbook(filename="../../../data/excel/Book.xlsx")

# sheetnames = workbook.sheetnames
# print('Sheetnames: ', sheetnames)

# sheet = workbook.active
# print(sheet)

# print(sheet.title)
# print(sheet['A1:A5'])

# headings = []
# for heading in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
#   print(heading)

# print(sheet['B6'].value)

# print(headings)

class Excel:
  filename: str
  workbook: Workbook
  sheet: Worksheet

  def __init__(self, filename, sheet = ''):
    workbook = load_workbook(filename=filename)

    self.filename = filename
    self.workbook = workbook

    print(sheet)
    active_sheet = ''
    for i, sheet_name in enumerate(workbook.worksheets):
      if sheet in workbook.worksheets:
        active_sheet = workbook.worksheets[i]
      # else:
      #   active_sheet = workbook.active
    if sheet not in workbook.worksheets:
      active_sheet = workbook.active

    self.sheet = active_sheet

  def sheetnames(self):
    return self.workbook.sheetnames

  def dimensions(self):
    return self.sheet.calculate_dimension()

  def get_sheet(self):
    return self.sheet

  def first_col(self):
    return re.search(r'^[A-Z]', self.dimensions()).group()

  def last_col(self):
    return re.search(r'(?<=:)[A-Z]', self.dimensions()).group()

  def first_row(self):
    return re.search(r'[0-9]', self.dimensions()).group()

  def last_row(self):
    return re.search(r'(?<=:[A-Z])[0-9]', self.dimensions()).group()

  def headings(self):
    # return self.sheet[f'{self.first_row()}:{self.last_row()}']
    heading_list = []
    for headings in self.sheet.iter_rows(min_row=1, max_row=1, values_only=True):
      # Returns tuple
      for h in headings:
        heading_list.append(h)

    print(self.sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    return heading_list


  def build_formula(self, formula):
    pass

  def get_formula(self, cell = ''):
    if cell != '':
      token = Tokenizer(self.sheet[cell].value)
      return token
    else:
      token = Tokenizer(self.sheet['B6'].value)
      # for t in token.items:
      #   print(t.value)

      # return token
      return token.formula

  def find_formulas(self):
    all_rows = self.sheet[f'{self.dimensions()}']

    formulas = []

    for row in all_rows:
      for cell in row:
        if re.search(r'=', str(cell.value)):
          # Formula label
          col_before = self.sheet[f'{self._num_to_abc_(cell.column - 1)}{cell.row}']
          formula = {
            'label':  f'Row: {cell.row}, Col: {cell.column}',
            'formula': cell.value
          }
          formulas.append(formula)

    return formulas

  def _num_to_abc_(self, num):
    data =  {
      1: 'A',
      2: 'B',
      3: 'C',
      4: 'D',
      5: 'E',
      6: 'F',
      7: 'G',
      8: 'H',
      9: 'I',
      10: 'J',
      11: 'K',
      12: 'L',
      13: 'M',
      14: 'N',
      15: 'O',
      16: 'P',
      17: 'Q',
      18: 'R',
      19: 'S',
      20: 'T',
      21: 'U',
      22: 'V',
      23: 'W',
      24: 'X',
      25: 'Y',
      26: 'Z'
    }

    return data[num]
  def parse_formula(self, formula):
    func = formulas.Parser().ast(formula)[1].compile()
    return func.inputs

# workbook = Excel(filename="../../../data/excel/hs-form.xlsx")

# workbook.save('xlsx-to-pdf.pdf', SaveFormat.PDF)
# import pandas as pd
# import pdfkit

# df = pd.read_excel("hs-form.xlsx")
# df.to_html("file.html")
# pdfkit.from_file("file.html", "file.pdf")

# workbook = load_workbook('fruits2.xlsx', guess_types=True, data_only=True)
# worksheet = workbook.active

# pw = PDFWriter('fruits2.pdf')
# pw.setFont('Courier', 12)
# pw.setHeader('XLSXtoPDF.py - convert XLSX data to PDF')
# pw.setFooter('Generated using openpyxl and xtopdf')

# ws_range = worksheet.iter_rows('A1:H13')
# for row in ws_range:
#     s = ''
#     for cell in row:
#         if cell.value is None:
#             s += ' ' * 11
#         else:
#             s += str(cell.value).rjust(10) + ' '
#     pw.writeLine(s)
# pw.savePage()
# pw.close()
# from pdfrw import pdfreader

# def get_pdf_info(path):
#   pdf = pdfreader(path)

#   print(pdf.keys())
#   print(pdf.info)
#   print(pdf.root.keys())
#   print('pdf has {} pages'.format(len(pdf.pages)))

# get_pdf_info('./hs-form.xlsx')
# formula_list = workbook.find_formulas()
# print(workbook.sheet)
# print(formula_list)
# print(workbook.parse_formula(formula_list[0]['formula']))
# f = open('../../../data/excel/formulas.json', 'w+')
# f.write(json.dumps(formula_list))
# f.close()
# print(formula_list)
# print('complete')